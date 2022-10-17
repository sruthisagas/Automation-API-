import openpyxl


class Test:

    def test_login_input_file_excel(self):
        input_file = openpyxl.load_workbook("C:\\Users\\user\\Sruthi\\TestData\\login_input.xlsx")
        user_data_sheet = input_file.active
        list_username = []
        list_password = []
        input_dict = {}
        for i in range(1, user_data_sheet.max_row + 1):  # to get rows
            list_username.append(user_data_sheet.cell(row=i, column=1).value)

        for j in range(1, user_data_sheet.max_row + 1):  # to get rows
            list_password.append(user_data_sheet.cell(row=j, column=2).value)

        return list_username, list_password
    
    def test_login_code_input_file_excel(self):
        input_file = openpyxl.load_workbook("C:\\Users\\user\\Sruthi\\TestData\\login_code_input.xlsx")
        user_data_sheet = input_file.active
        list_username = []
        for i in range(1, user_data_sheet.max_row + 1):  # to get rows
            list_username.append(user_data_sheet.cell(row=i, column=1).value)

        return list_username

    def test_auth_token_data_file_excel(self):
        input_file = openpyxl.load_workbook("C:\\Users\\user\\Sruthi\\TestData\\user_data_update_input.xlsx")
        user_data_sheet = input_file.active
        list_auth_token = []
        for i in range(1, user_data_sheet.max_row + 1):  # to get rows
            list_auth_token.append(user_data_sheet.cell(row=i, column=1).value)

        return list_auth_token

    def test_send_confirmation_code_file_excel(self):
        input_file = openpyxl.load_workbook("C:\\Users\\user\\Sruthi\\TestData\\"
                                            "test_send_confirmation_code_to_email_input.xlsx")
        user_data_sheet = input_file.active
        list_username_code = []
        for i in range(1, user_data_sheet.max_row + 1):  # to get rows
            list_username_code.append(user_data_sheet.cell(row=i, column=1).value)

        return list_username_code

    def test_reset_password_file_excel(self):
        input_file = openpyxl.load_workbook("C:\\Users\\user\\Sruthi\\TestData\\"
                                            "reset_password_input.xlsx")
        user_data_sheet = input_file.active
        list_username_reset_password = []
        for i in range(1, user_data_sheet.max_row + 1):  # to get rows
            list_username_reset_password.append(user_data_sheet.cell(row=i, column=1).value)

        return list_username_reset_password

    def test_verify_code_file_excel(self):
        input_file = openpyxl.load_workbook("C:\\Users\\user\\Sruthi\\TestData\\"
                                            "verify_code_input.xlsx")
        user_data_sheet = input_file.active
        list_username_verify_code = []
        for i in range(1, user_data_sheet.max_row + 1):  # to get rows
            list_username_verify_code.append(user_data_sheet.cell(row=i, column=1).value)

        return list_username_verify_code

    def test_input_file_excel(self):
        input_file = openpyxl.load_workbook("C:\\Users\\user\\Sruthi\\TestData\\registration_input.xlsx")
        user_data_sheet = input_file.active
        list_username = []
        list_password = []
        input_dict = {}
        for i in range(1, user_data_sheet.max_row + 1):  # to get rows
            list_username.append(user_data_sheet.cell(row=i, column=1).value)

        for j in range(1, user_data_sheet.max_row + 1):  # to get rows
            list_password.append(user_data_sheet.cell(row=j, column=2).value)

        '''k = 0
        while k < len(list_username):
            input_dict[list_username[k]] = list_password[k]
            k += 1'''

        return list_username, list_password
